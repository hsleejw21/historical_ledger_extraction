"""
pipeline/run_pipeline.py

Production pipeline for historical ledger extraction (v2_no_claude).

Architecture:
    Image → gemini-flash extractor ─┐
                                     ├─→ gemini-flash supervisor → Excel output
    Image → gpt-5-mini extractor  ──┘

Each input image produces one sheet in the output Excel file.
Intermediate extractor/supervisor JSONs are cached in pipeline/cache/
so re-runs skip API calls for already-processed pages.

Usage:
    # Process a directory of images:
    python pipeline/run_pipeline.py --images data/images/

    # Process specific files:
    python pipeline/run_pipeline.py --images data/images/1700_7.png data/images/1700_8.png

    # Reuse cached intermediate results:
    python pipeline/run_pipeline.py --images data/images/ --use-cache

    # Specify output path:
    python pipeline/run_pipeline.py --images data/images/ --output results/extraction.xlsx
"""
import os
import re
import sys
import json
import argparse
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Allow running from any working directory
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.agents.standalone_extractor import StandaloneExtractor
from src.agents.supervisor import Supervisor

# ---------------------------------------------------------------------------
# Pipeline configuration (v2_no_claude)
# ---------------------------------------------------------------------------
EXTRACTOR_MODELS = ["gemini-flash", "gpt-5-mini"]
SUPERVISOR_MODEL = "gemini-flash"

# ---------------------------------------------------------------------------
# Output columns (clean final data — no internal debug fields)
# ---------------------------------------------------------------------------
OUTPUT_COLUMNS = [
    "row_index",
    "row_type",
    "description",
    "amount_pounds",
    "amount_shillings",
    "amount_pence_whole",
    "amount_pence_fraction",
]

HEADER_LABELS = {
    "row_index":            "Row #",
    "row_type":             "Type",
    "description":          "Description",
    "amount_pounds":        "£ (Pounds)",
    "amount_shillings":     "s (Shillings)",
    "amount_pence_whole":   "d (Pence)",
    "amount_pence_fraction":"d Fraction",
}

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
ROW_TYPE_FILLS = {
    "header": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
    "total":  PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "entry":  None,
}
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

COL_WIDTHS = {
    "row_index":            6,
    "row_type":             10,
    "description":          45,
    "amount_pounds":        12,
    "amount_shillings":     14,
    "amount_pence_whole":   12,
    "amount_pence_fraction":12,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _natural_sort_key(s: str) -> list:
    return [int(c) if c.isdigit() else c for c in re.split(r"(\d+)", s)]


def _discover_images(path_args: list) -> dict:
    """Accept file paths and/or directories; return {page_name: image_path}."""
    images = {}
    for arg in path_args:
        if os.path.isdir(arg):
            for fname in sorted(os.listdir(arg)):
                if fname.lower().endswith((".png", ".jpg", ".jpeg")):
                    stem = os.path.splitext(fname)[0]
                    images[stem] = os.path.join(arg, fname)
        elif os.path.isfile(arg):
            stem = os.path.splitext(os.path.basename(arg))[0]
            images[stem] = arg
        else:
            print(f"[Warning] Path not found: {arg}")
    return images


def _cache_path(cache_dir: str, page_name: str, label: str) -> str:
    return os.path.join(cache_dir, f"{page_name}_{label}.json")


def _load_json(path: str):
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None
    return None


def _save_json(path: str, data: dict):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# ---------------------------------------------------------------------------
# Per-page extraction
# ---------------------------------------------------------------------------
def _process_page(image_path: str, page_name: str, cache_dir: str, use_cache: bool) -> list:
    """Run extractors + supervisor for one page. Returns the final rows list."""
    candidates = {}

    for model_key in EXTRACTOR_MODELS:
        cache_file = _cache_path(cache_dir, page_name, f"extractor_{model_key}")
        cached = _load_json(cache_file) if use_cache else None

        if cached is not None:
            print(f"    [cache]  extractor/{model_key}")
            out = cached
        else:
            print(f"    [run]    extractor/{model_key} ...")
            out = StandaloneExtractor(model_key=model_key).run(image_path)
            _save_json(cache_file, out)

        if out:
            candidates[model_key] = out

    if not candidates:
        print(f"    [error]  No extractor output — skipping page")
        return []

    sup_cache = _cache_path(cache_dir, page_name, f"supervisor_{SUPERVISOR_MODEL}")
    cached_sup = _load_json(sup_cache) if use_cache else None

    if cached_sup is not None:
        print(f"    [cache]  supervisor/{SUPERVISOR_MODEL}")
        sup_out = cached_sup
    else:
        print(f"    [run]    supervisor/{SUPERVISOR_MODEL} ({len(candidates)} candidates) ...")
        sup_out = Supervisor(model_key=SUPERVISOR_MODEL).run(image_path, candidates)
        _save_json(sup_cache, sup_out)

    return sup_out.get("rows", [])


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------
def _rows_to_df(rows: list) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=[HEADER_LABELS[c] for c in OUTPUT_COLUMNS])
    records = [
        {HEADER_LABELS[col]: row.get(col, "") for col in OUTPUT_COLUMNS}
        for row in rows
    ]
    return pd.DataFrame(records)


def _style_sheet(ws, n_rows: int):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        row_type = str(row[1].value or "").lower()
        fill = ROW_TYPE_FILLS.get(row_type)
        for cell in row:
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top")

    for i, col_key in enumerate(OUTPUT_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(col_key, 15)

    ws.freeze_panes = "A2"


def _write_summary(writer, page_stats: list):
    df = pd.DataFrame(
        page_stats,
        columns=["Page", "Total Rows", "Entries", "Headers", "Totals"]
    )
    df.to_excel(writer, sheet_name="Summary", index=False)
    ws = writer.sheets["Summary"]
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for i, w in enumerate([22, 12, 12, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------
def run_pipeline(image_paths: list, output_path: str, cache_dir: str, use_cache: bool):
    images = _discover_images(image_paths)
    if not images:
        print("[Error] No images found.")
        return

    sorted_pages = sorted(images.keys(), key=_natural_sort_key)

    print(f"\n{'='*60}")
    print(f"  Historical Ledger Extraction Pipeline (v2_no_claude)")
    print(f"{'='*60}")
    print(f"  Extractors : {', '.join(EXTRACTOR_MODELS)}")
    print(f"  Supervisor : {SUPERVISOR_MODEL}")
    print(f"  Pages      : {len(sorted_pages)}")
    print(f"  Cache dir  : {cache_dir}")
    print(f"  Output     : {output_path}")
    print(f"{'='*60}\n")

    os.makedirs(cache_dir, exist_ok=True)
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    page_stats = []

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Placeholder so openpyxl creates the workbook
        pd.DataFrame().to_excel(writer, sheet_name="Summary", index=False)

        for page_name in sorted_pages:
            image_path = images[page_name]
            print(f"[Page] {page_name}")

            try:
                rows = _process_page(image_path, page_name, cache_dir, use_cache)
            except Exception as e:
                print(f"    [error] {e}")
                rows = []

            df = _rows_to_df(rows)
            sheet_name = page_name[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_sheet(writer.sheets[sheet_name], len(rows))

            page_stats.append([
                page_name,
                len(rows),
                sum(1 for r in rows if str(r.get("row_type", "")).lower() == "entry"),
                sum(1 for r in rows if str(r.get("row_type", "")).lower() == "header"),
                sum(1 for r in rows if str(r.get("row_type", "")).lower() == "total"),
            ])
            print(f"    -> {len(rows)} rows written to sheet '{sheet_name}'\n")

        _write_summary(writer, page_stats)
        wb = writer.book
        wb.move_sheet("Summary", offset=-len(sorted_pages))

    print(f"{'='*60}")
    print(f"  Done. Results saved to:")
    print(f"  {output_path}")
    print(f"{'='*60}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    pipeline_dir = os.path.dirname(os.path.abspath(__file__))

    parser = argparse.ArgumentParser(
        description="Historical Ledger Extraction — v2_no_claude production pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python pipeline/run_pipeline.py --images data/images/
  python pipeline/run_pipeline.py --images data/images/1700_7.png
  python pipeline/run_pipeline.py --images data/images/ --use-cache
  python pipeline/run_pipeline.py --images data/images/ --output results/ledger.xlsx
        """,
    )
    parser.add_argument(
        "--images", nargs="+", required=True,
        metavar="PATH",
        help="Image files (.png/.jpg) or directories to process.",
    )
    parser.add_argument(
        "--output", default=None,
        metavar="FILE",
        help=(
            "Output Excel file path. "
            "Default: pipeline/output/ledger_<YYYYMMDD_HHMMSS>.xlsx"
        ),
    )
    parser.add_argument(
        "--cache-dir", default=os.path.join(pipeline_dir, "cache"),
        metavar="DIR",
        help="Directory for intermediate JSON results (default: pipeline/cache/).",
    )
    parser.add_argument(
        "--use-cache", action="store_true", default=False,
        help="Reuse cached extractor/supervisor results (skip API calls).",
    )
    args = parser.parse_args()

    if args.output is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = os.path.join(pipeline_dir, "output", f"ledger_{timestamp}.xlsx")

    run_pipeline(args.images, args.output, args.cache_dir, args.use_cache)


if __name__ == "__main__":
    main()
