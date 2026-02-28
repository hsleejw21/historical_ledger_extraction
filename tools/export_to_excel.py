"""
tools/export_to_excel.py

Convert existing sample_pdf supervisor results (JSON) into a single Excel file
with one sheet per page.

Usage:
    python tools/export_to_excel.py
    python tools/export_to_excel.py --results-dir experiments/results/sample_pdf
    python tools/export_to_excel.py --output my_results.xlsx
"""
import os
import re
import sys
import json
import argparse

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_RESULTS_DIR = os.path.join(BASE_DIR, "experiments", "results", "sample_pdf")
DEFAULT_OUTPUT = os.path.join(BASE_DIR, "experiments", "results", "sample_pdf_results.xlsx")

# Columns written to each sheet (in order)
DATA_COLUMNS = [
    "row_index",
    "row_type",
    "description",
    "amount_pounds",
    "amount_shillings",
    "amount_pence_whole",
    "amount_pence_fraction",
    "confidence_score",
    "notes",
]

# Friendly header labels
HEADER_LABELS = {
    "row_index":            "Row #",
    "row_type":             "Type",
    "description":          "Description",
    "amount_pounds":        "£ (Pounds)",
    "amount_shillings":     "s (Shillings)",
    "amount_pence_whole":   "d (Pence)",
    "amount_pence_fraction":"d Fraction",
    "confidence_score":     "Confidence",
    "notes":                "Supervisor Notes",
}

# Row-type fill colours (light pastel)
ROW_TYPE_FILLS = {
    "header": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),  # blue-grey
    "total":  PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # green
    "entry":  None,
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _natural_sort_key(s: str) -> list:
    """Sort page names like 1700_1, 1700_2, …, 1700_10, 1704_1, …"""
    return [int(c) if c.isdigit() else c for c in re.split(r"(\d+)", s)]


def load_supervisor_results(results_dir: str) -> dict:
    """Return {page_name: data} for every supervisor JSON in results_dir."""
    results = {}
    for fname in os.listdir(results_dir):
        if "_supervisor_" not in fname or not fname.endswith(".json"):
            continue
        page_name = fname.split("_supervisor_")[0]
        path = os.path.join(results_dir, fname)
        try:
            with open(path, "r", encoding="utf-8") as f:
                results[page_name] = json.load(f)
        except Exception as e:
            print(f"  [Warning] Could not load {fname}: {e}")
    return results


def _style_sheet(ws, n_rows: int):
    """Apply header formatting and column widths to a worksheet."""
    # Style header row
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        row_type = str(row[1].value or "").lower()  # column B = row_type
        fill = ROW_TYPE_FILLS.get(row_type)
        for cell in row:
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column == len(DATA_COLUMNS)))

    # Column widths
    col_widths = {
        "row_index":            6,
        "row_type":             10,
        "description":          40,
        "amount_pounds":        12,
        "amount_shillings":     14,
        "amount_pence_whole":   12,
        "amount_pence_fraction":12,
        "confidence_score":     12,
        "notes":                60,
    }
    for i, col_key in enumerate(DATA_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col_key, 15)

    # Freeze the header row
    ws.freeze_panes = "A2"


def _make_summary_sheet(writer, page_stats: list):
    """Write a Summary sheet with page-level row counts."""
    df = pd.DataFrame(page_stats, columns=["Page", "Total Rows", "Entries", "Headers", "Totals"])
    df.to_excel(writer, sheet_name="Summary", index=False)
    ws = writer.sheets["Summary"]
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for col, width in zip("ABCDE", [20, 12, 12, 12, 12]):
        ws.column_dimensions[get_column_letter(ord(col) - ord("A") + 1)].width = width
    ws.freeze_panes = "A2"


def export_to_excel(results_dir: str, output_path: str):
    print(f"Loading results from: {results_dir}")
    results = load_supervisor_results(results_dir)

    if not results:
        print("[Error] No supervisor result files found.")
        return

    sorted_pages = sorted(results.keys(), key=_natural_sort_key)
    print(f"Found {len(sorted_pages)} pages\n")

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    page_stats = []

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Write Summary sheet first (placeholder — will be overwritten below)
        pd.DataFrame().to_excel(writer, sheet_name="Summary", index=False)

        for page_name in sorted_pages:
            data = results[page_name]
            rows = data.get("rows", [])

            # Build records
            records = [{HEADER_LABELS[col]: row.get(col, "") for col in DATA_COLUMNS} for row in rows]
            df = pd.DataFrame(records, columns=[HEADER_LABELS[c] for c in DATA_COLUMNS])

            sheet_name = page_name[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_sheet(writer.sheets[sheet_name], len(rows))

            counts = {r.get("row_type", "").lower() for r in rows}
            page_stats.append([
                page_name,
                len(rows),
                sum(1 for r in rows if r.get("row_type", "").lower() == "entry"),
                sum(1 for r in rows if r.get("row_type", "").lower() == "header"),
                sum(1 for r in rows if r.get("row_type", "").lower() == "total"),
            ])
            print(f"  [Sheet] {sheet_name:<25} {len(rows):>3} rows")

        # Overwrite Summary with real data
        _make_summary_sheet(writer, page_stats)

        # Move Summary to the first position
        wb = writer.book
        wb.move_sheet("Summary", offset=-len(sorted_pages))

    print(f"\n[Done] Excel saved to: {output_path}")
    print(f"       Sheets: Summary + {len(sorted_pages)} page sheets")


def main():
    parser = argparse.ArgumentParser(
        description="Export sample_pdf supervisor results to a single Excel file."
    )
    parser.add_argument(
        "--results-dir", default=DEFAULT_RESULTS_DIR,
        help=f"Directory with supervisor JSON files (default: {DEFAULT_RESULTS_DIR})"
    )
    parser.add_argument(
        "--output", default=DEFAULT_OUTPUT,
        help=f"Output .xlsx path (default: {DEFAULT_OUTPUT})"
    )
    args = parser.parse_args()
    export_to_excel(args.results_dir, args.output)


if __name__ == "__main__":
    main()
